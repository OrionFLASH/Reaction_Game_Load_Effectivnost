#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Программа для обработки данных из Excel файлов
Автор: AI Assistant
Версия: 1.0.0
Дата создания: 2024
"""

import os
import sys
import time
import logging
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from pathlib import Path
import traceback

# =============================================================================
# КОНСТАНТЫ И НАСТРОЙКИ ПРОГРАММЫ
# =============================================================================

# Путь к рабочей папке (сырая строка для избежания экранирования)
WORK_DIR = r"/Users/orionflash/Desktop/MyProject/Reaction_Effectiv_LOAD/WORK"

# Названия подпапок
INPUT_FOLDER = "INPUT"      # Папка с входными файлами
OUTPUT_FOLDER = "OUTPUT"    # Папка с выходными файлами
LOGS_FOLDER = "LOGS"        # Папка с логами

# Настройки входных файлов (имя без расширения, расширение отдельно)
INPUT_FILES = [
    {"name": "data1_20250822_153515", "extension": ".xlsx"},
    {"name": "data2_20250822_153515", "extension": ".xlsx"}
]

# Настройки выходных файлов
OUTPUT_FILES = [
    {"name": "processed_data", "extension": ".xlsx", "suffix_format": "_YYYYMMDD-HHMMSS"}
    # ЗАКЛАДКА ДЛЯ БУДУЩЕГО CSV:
    # {"name": "processed_data", "extension": ".csv", "suffix_format": "_YYYYMMDD-HHMMSS"}
]

# Настройки лог-файла
LOG_FILE = {
    "name": "processing_log",
    "extension": ".log",
    "suffix_format": "_YYYYMMDD"
}

# Режим работы программы
# "process" - обработка данных (основная работа)
# "create-test" - создание тестовых данных
PROGRAM_MODE = "process"
#PROGRAM_MODE = "create-test"

# Уровень логирования (INFO или DEBUG)
LOG_LEVEL = "DEBUG"

# Параметры генерации тестовых данных
DATA_PARAMS = {
    "total_employees": 1600,        # Общее количество сотрудников
    "effective_share": 0.80,        # Доля эффективных сотрудников (80%)
    "operational_income_final_min": 500000,    # Минимальный операционный доход на 31 июля 2025 (финальный период, тыс. руб.)
    "operational_income_final_max": 220000000,   # Максимальный операционный доход на 31 июля 2025 (финальный период, тыс. руб.)
    "operational_income_current_min": 500000,  # Минимальный операционный доход на 20 августа 2025 (текущий период, тыс. руб.)
    "operational_income_current_max": 220000000, # Максимальный операционный доход на 20 августа 2025 (текущий период, тыс. руб.)
    "employee_overlap": 0.90,       # Доля одинаковых сотрудников в двух файлах (90%)
    "new_employees_share": 0.05,    # Доля новых сотрудников (5%)
    "removed_employees_share": 0.05 # Доля убранных сотрудников (5%)
}

# Процентили для ранжирования (25%, 50%, 75%)
PERCENTILES = [25, 50, 75]

# Настройки форматирования колонок Excel
# Универсальная система управления форматированием через параметры
# 
# ПАРАМЕТРЫ ФОРМАТИРОВАНИЯ:
# - 'format': тип данных ('text', 'number', 'date')
# - 'width': ширина колонки в символах
# - 'alignment': выравнивание ('left', 'center', 'right')
# - 'number_format': числовой формат (только для format='number')
#
# ЧИСЛОВЫЕ ФОРМАТЫ:
# - '0' - целое число
# - '0.0' - число с 1 знаком после запятой
# - '#,##0.0' - число с разделителями разрядов и 1 знаком после запятой
# - '#,##0' - целое число с разделителями разрядов
# - '0.00%' - процент с 2 знаками после запятой
#
# ДАТЫ:
# - 'DD.MM.YYYY' - день.месяц.год
# - 'MM/DD/YYYY' - месяц/день/год
# - 'YYYY-MM-DD' - год-месяц-день
#
# ПРИМЕРЫ ИСПОЛЬЗОВАНИЯ:
# - Финансовые данные: {'format': 'number', 'number_format': '#,##0.0', 'alignment': 'right'}
# - Проценты: {'format': 'number', 'number_format': '0.0%', 'alignment': 'center'}
# - Даты: {'format': 'date', 'number_format': 'DD.MM.YYYY', 'alignment': 'center'}
# - Текст: {'format': 'text', 'alignment': 'left'}
# Группы форматирования колонок
COLUMN_FORMAT_GROUPS = {
    # ТЕКСТОВЫЕ КОЛОНКИ (левое выравнивание)
    'text_left': {
        'columns': ['ТБ', 'ГОСБ', 'ФИО', 'вывод'],
        'format': 'text',
        'width': 20,
        'alignment': 'left'
    },
    
    # ЦЕНТРИРОВАННЫЕ КОЛОНКИ
    'text_center': {
        'columns': ['ЭФ.КМ', 'вып условий', 'КОД вывода'],
        'format': 'text', 
        'width': 12,
        'alignment': 'center'
    },
    
    # ФИНАНСОВЫЕ ДАННЫЕ (с разделителями разрядов, 1 знак после запятой)
    'financial': {
        'columns': ['ОД ТЕКУЩИЙ', 'ОД ПРОШЛЫЙ', 'прирост'],
        'format': 'number',
        'number_format': '#,##0.0',
        'width': 18,
        'alignment': 'right'
    },
    
    # ПРОЦЕНТИЛИ (с разделителями разрядов, 1 знак после запятой)
    'percentiles': {
        'columns': ['СТРАНА 50', 'СТРАНА 75', 'СТРАНА 90', 'ТБ 25', 'ТБ 50', 'ТБ 75', 'ГОСБ 25', 'ГОСБ 50', 'ГОСБ 75'],
        'format': 'number',
        'number_format': '#,##0.0',
        'width': 15,
        'alignment': 'right'
    },
    
    # РАНГИ (1 знак после запятой)
    'ranks': {
        'columns': ['ранг ОД BANK', 'ранг ОД TB', 'темп'],
        'format': 'number',
        'number_format': '0.0',
        'width': 15,
        'alignment': 'right'
    },
    
    # ЦЕЛЫЕ ЧИСЛА
    'integers': {
        'columns': ['число страна', 'число ТБ', 'число подразделение'],
        'format': 'number',
        'number_format': '0',
        'width': 15,
        'alignment': 'right'
    }
}

# Специальные настройки для отдельных колонок (переопределяют групповые)
# 
# СПЕЦИАЛЬНЫЕ ТИПЫ ФОРМАТИРОВАНИЯ:
# - 'padded_number': число с лидирующими нулями до заданного количества знаков
#   - 'total_digits': общее количество знаков (например, 10)
#   - 'leading_zeros': количество лидирующих нулей (обычно равно total_digits)
#   - 'min_value': минимальное значение для генерации
#   - 'max_value': максимальное значение для генерации
#
# ПРИМЕРЫ НАСТРОЙКИ ТН 10:
# - 10 знаков, всегда заполнено нулями: {'total_digits': 10, 'leading_zeros': 10}
# - 8 знаков, минимум 2 нуля: {'total_digits': 8, 'leading_zeros': 2}
# - 12 знаков, без ограничений: {'total_digits': 12, 'leading_zeros': 0}
#
COLUMN_SPECIAL_FORMATS = {
    'ТН 10': {
        'width': 12,              # Ширина колонки
        'format_type': 'padded_number',  # Специальный тип форматирования
        'total_digits': 10,       # Общее количество знаков
        'leading_zeros': 10,      # Количество лидирующих нулей (всегда заполняем до 10 знаков)
        'min_value': 1,           # Минимальное значение
        'max_value': 9999999999   # Максимальное значение (10 знаков)
    },
    'ТБ': {'width': 20},         # Названия банков
    'ГОСБ': {'width': 30},       # Названия отделений
    'ФИО': {'width': 25},        # ФИО сотрудников
    'вывод': {'width': 40}       # Текстовые описания
}

# Структура территориальных банков и их головных отделений
# Вложенный словарь: ТБ -> список ГОСБ
BANK_STRUCTURE = {
    "Байкальский банк": [
        "Аппарат Байкальского Банка",
        "Бурятское ГОСБ №8601",
        "Иркутское ГОСБ №8586",
        "Читинское ГОСБ №8600",
        "Якутское ГОСБ №8603"
    ],
    
    "Волго-Вятский банк": [
        "Аппарат Волго-Вятского банка",
        "Банк Татарстан ГОСБ №8610",
        "Владимирское ГОСБ №8611",
        "Головное отделение по Нижегородской области",
        "Кировское ГОСБ №8612",
        "Марий Эл ГОСБ №8614",
        "Мордовское ГОСБ №8589",
        "Пермское ГОСБ №6984",
        "Удмуртское ГОСБ №8618",
        "Чувашское ГОСБ №8613"
    ],
    
    "Дальневосточный банк": [
        "Аппарат Дальневосточного банка",
        "Биробиджанское ГОСБ №4157",
        "Благовещенское ГОСБ №8636",
        "Камчатское ГОСБ №8556",
        "Приморское ГОСБ №8635",
        "Северо-Восточное ГОСБ №8645",
        "Хабаровское ГОСБ №9070",
        "Чукотское ГОСБ №8557",
        "Южно-Сахалинское ГОСБ №8567"
    ],
    
    "Московский банк": [
        "УПРАВЛЕНИЕ по работе с предприятиями инфраструктуры",
        "УПРАВЛЕНИЕ по работе с предприятиями промышленности",
        "УПРАВЛЕНИЕ по работе с предприятиями сферы недвижимости",
        "УПРАВЛЕНИЕ по работе с предприятиями сферы услуг",
        "УПРАВЛЕНИЕ по работе с предприятиями торговли"
    ],
    
    "Поволжский банк": [
        "Аппарат Поволжского банка",
        "Астраханское ГОСБ №8625",
        "Волгоградское ГОСБ №8621",
        "Оренбургское ГОСБ №8623",
        "Пензенское ГОСБ №8624",
        "Самарское ГОСБ №6991",
        "Саратовское ГОСБ №8622",
        "Ульяновское ГОСБ №8588"
    ],
    
    "Северо-Западный банк": [
        "Аппарат Северо-Западного банка",
        "Архангельское ГОСБ №8637",
        "Вологодское ГОСБ №8638",
        "ГО по г. Санкт-Петербургу",
        "ГО по Ленинградской области",
        "Калининградское ГОСБ №8626",
        "Карельское ГОСБ №8628",
        "Коми ГОСБ №8617",
        "Мурманское ГОСБ №8627",
        "Новгородское ГОСБ №8629",
        "Псковское ГОСБ №8630"
    ],
    
    "Сибирский банк": [
        "Аппарат Сибирского банка",
        "Абаканское ГОСБ №8602",
        "Алтайское ГОСБ №8644",
        "Кемеровское ГОСБ №8615",
        "Красноярское ГОСБ №8646",
        "Новосибирское ГОСБ №8047",
        "Омское ГОСБ №8634",
        "Томское ГОСБ №8616"
    ],
    
    "Среднерусский банк": [
        "Аппарат Среднерусского банка",
        "Брянское ГОСБ №8605",
        "Восточное ГОСБ №1023",
        "Западное ГОСБ №1025",
        "Ивановское ГОСБ №8639",
        "Калужское ГОСБ №8608",
        "Костромское ГОСБ №8640",
        "Рязанское ГОСБ №8606",
        "Северное ГОСБ №1026",
        "Смоленское ГОСБ №8609",
        "Тверское ГОСБ №8607",
        "Тульское ГОСБ №8604",
        "Южное ГОСБ №1024",
        "Ярославское ГОСБ №17"
    ],
    
    "Уральский банк": [
        "Аппарат Уральского банка",
        "Башкирское ГОСБ №8598",
        "Западно-Сибирское ГОСБ №8647",
        "Курганское ГОСБ №8599",
        "Свердловское ГОСБ №7003",
        "Челябинское ГОСБ №8597",
        "Югорское отделение №5940",
        "Ямало-Ненецкое отделение №8369"
    ],
    
    "Центрально-Черноземный банк": [
        "Аппарат Центрально-Черноземного банка",
        "Белгородское ГОСБ №8592",
        "ГО по Воронежской области",
        "Головное отделение по ЛНР",
        "Курское ГОСБ №8596",
        "Липецкое ГОСБ №8593",
        "Орловское ГОСБ №8595",
        "Тамбовское ГОСБ №8594"
    ],
    
    "Юго-Западный банк": [
        "Аппарат Юго-Западного банка",
        "Адыгейское ОСБ №8620",
        "Головное отделение по ДНР",
        "Головное отделение по Республике Крым",
        "Ингушское отделение № 8633",
        "Кабардино-Балкарское ОСБ №8631",
        "Калмыцкое ОСБ №8579",
        "Карачаево-Черкесское ОСБ №8585",
        "Краснодарское ГОСБ №8619",
        "Ростовское ГОСБ №5221",
        "Северо-Осетинское ОСБ №8632",
        "Ставропольское ГОСБ №5230",
        "Чеченское ОСБ №8643"
    ]
}

# Вспомогательные списки для обратной совместимости
TERRITORIAL_BANKS = list(BANK_STRUCTURE.keys())
HEAD_OFFICES = [gosb for gosb_list in BANK_STRUCTURE.values() for gosb in gosb_list]

# Сообщения для логирования
LOG_MESSAGES = {
    "start": "Программа запущена",
    "end": "Программа завершена",
    "processing_start": "Начало обработки данных",
    "processing_end": "Обработка данных завершена",
    "file_loaded": "Файл {} загружен успешно",
    "file_saved": "Файл {} сохранен успешно",
    "error": "Ошибка: {}",
    "summary": "Сводка выполнения: {}",
    "time_elapsed": "Время выполнения: {}",
    "files_processed": "Обработано файлов: {}",
    "outputs_created": "Создано выходных файлов: {}",
    "errors_count": "Количество ошибок: {}",
    "test_data_created": "Тестовые данные созданы успешно",
    "test_file_created": "Создан тестовый файл: {}",
    "test_data_info": "Тестовый файл содержит {} строк и {} столбцов",
    "data_generation_start": "Начало генерации тестовых данных",
    "data_generation_end": "Генерация тестовых данных завершена",
    "employees_created": "Создано сотрудников: {}",
    "tb_distribution": "Распределение по ТБ: {}",
    "gosb_distribution": "Распределение по ГОСБ: {}",
    "effective_distribution": "Распределение эффективных: {}",
    "directory_ready": "Директория {} готова к работе",
    "tb_mapping_created": "Создано распределение ГОСБ по ТБ: {} ТБ",
    "progress_employees": "Сгенерировано сотрудников: {}",
    "unique_tn_fio": "Уникальных ТН: {}, Уникальных ФИО: {}",
    "duplicate_tn_error": "ОШИБКА: Дублирование табельных номеров!",
    "duplicate_fio_error": "ОШИБКА: Дублирование ФИО!",
    "file_saved_debug": "Файл сохранен: {}",
    "file_size_debug": "Размер файла: {} строк, {} столбцов",
    "generation_error": "Ошибка при генерации тестовых данных: {}",
    "save_error": "Ошибка при сохранении файла: {}",
    "details_error": "Детали ошибки: {}",
    "file_not_found": "Файл {} не найден",
    "load_file_error": "Ошибка при загрузке файла {}: {}",
    "no_data_to_process": "Нет данных для обработки",
    "no_data_to_save": "Нет данных для сохранения",
    "data_processed": "Обработано данных: {} строк",
    "rows_columns_loaded": "Загружено {} строк и {} столбцов из файла {}",
    "duplicates_removed": "Удалено дубликатов: {}",
    "missing_values_filled": "Заполнено пропущенных значений: {}",
    "processing_error": "Ошибка при обработке данных: {}",
    "critical_error": "Критическая ошибка в процессе выполнения: {}",
    "mode_create_test": "Режим: Создание тестовых данных",
    "mode_process": "Режим: Обработка данных",
    "test_data_success": "Тестовые данные созданы успешно. Проверьте папку INPUT.",
    "process_success": "Программа выполнена успешно. Проверьте папку OUTPUT для результатов.",
    "main_critical_error": "Критическая ошибка при запуске программы: {}",
    "file_saved_debug_old": "Файл сохранен: {}",
    "analysis_file1": "=== Анализ файла 1 (31 июля 2025 года) ===",
    "analysis_file2": "=== Анализ файла 2 (20 августа 2025 года) ===",
    "analysis_overlap": "=== Анализ перекрытия сотрудников ===",
    "employees_2024": "Сотрудников в 2024: {}",
    "employees_2025": "Сотрудников в 2025: {}",
    "overlap_info": "Перекрытие (одинаковые): {} ({:.1f}%)",
    "new_employees_info": "Новых сотрудников: {} ({:.1f}%)",
    "removed_employees_info": "Убранных сотрудников: {} ({:.1f}%)",
    "unique_tn_fio_file1": "Файл 1 - Уникальных ТН: {}, Уникальных ФИО: {}",
    "unique_tn_fio_file2": "Файл 2 - Уникальных ТН: {}, Уникальных ФИО: {}",
    "employees_july_count": "Сотрудников на 31 июля: {}",
    "employees_august_count": "Сотрудников на 20 августа: {}",
    "overlap_info_detailed": "Перекрытие (одинаковые): {} ({:.1f}%)",
    "new_employees_info_detailed": "Новых сотрудников: {} ({:.1f}%)",
    "removed_employees_info_detailed": "Убранных сотрудников: {} ({:.1f}%)",
    "unique_tn_fio_debug": "Файл {} - Уникальных ТН: {}, Уникальных ФИО: {}",
    "files_not_found_error": "Не найдены файлы {}.xlsx или {}.xlsx",
    "files_loaded_debug": "Загружены файлы: data1 ({} строк), data2 ({} строк)",
    "unique_tn_created": "Создан список из {} уникальных ТН",
    "ranks_calculation": "Рассчитываем ранги ОД...",
    "percentiles_calculation": "Рассчитываем процентили...",
    "ranking_calculation": "Рассчитываем ранжирование по процентилям...",
    "data_processed_debug": "Обработано данных: {} строк, {} колонок",
    "load_time_debug": "Время загрузки файлов: {}",
    "process_time_debug": "Время обработки данных: {}",
    "save_time_debug": "Время сохранения файлов: {}",
    "file1_info": "Файл 1 (31 июля) - Уникальных ТН: {}, Уникальных ФИО: {}",
    "file2_info": "Файл 2 (20 августа) - Уникальных ТН: {}, Уникальных ФИО: {}",
    "files_loaded_info": "Загружены файлы: data1 ({} строк), data2 ({} строк)",
    "unique_tn_list_created": "Создан список из {} уникальных ТН",
    "data_processing_time": "Время обработки данных: {}",
    "data_processed_info": "Обработано данных: {} строк, {} колонок",
    "file_loading_time": "Время загрузки файлов: {}",
    "file_saving_time": "Время сохранения файлов: {}",
    "files_not_found": "Не найдены файлы {}.xlsx или {}.xlsx",
    "autofilter_added": "Добавлен автофильтр на диапазон {}",
    "panes_frozen": "Установлена фиксация панелей на уровне A2 (заголовки видны)",
    "columns_formatted": "Отформатированы {} колонок по содержимому",
    "group_formatting_applied": "Применено групповое форматирование: {} групп, {} колонок",
    "special_formats_applied": "Специальные настройки применены к {} колонкам",
    "padded_number_formatted": "Применено специальное форматирование к {} колонкам (padded_number)"
}

# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

def format_execution_time(seconds):
    """
    Форматирует время выполнения в формат MM:SS.mmm
    
    Args:
        seconds (float): Время в секундах
        
    Returns:
        str: Время в формате MM:SS.mmm
    """
    if seconds < 60:
        return f"00:{seconds:06.3f}"
    else:
        minutes = int(seconds // 60)
        remaining_seconds = seconds % 60
        return f"{minutes:02d}:{remaining_seconds:06.3f}"

# =============================================================================
# КЛАСС ДЛЯ ЛОГИРОВАНИЯ
# =============================================================================

class DataProcessorLogger:
    """Класс для управления логированием программы"""
    
    def __init__(self, log_dir, log_name, log_extension, suffix_format, level="INFO"):
        """
        Инициализация логгера
        
        Args:
            log_dir (str): Директория для логов
            log_name (str): Имя лог-файла без расширения
            log_extension (str): Расширение лог-файла
            suffix_format (str): Формат суффикса для имени файла
            level (str): Уровень логирования (INFO/DEBUG)
        """
        self.log_dir = Path(log_dir)
        self.log_name = log_name
        self.log_extension = log_extension
        self.suffix_format = suffix_format
        self.level = level.upper()
        
        # Создаем директорию для логов если её нет
        self.log_dir.mkdir(parents=True, exist_ok=True)
        
        # Формируем имя лог-файла с суффиксом
        timestamp = datetime.now().strftime(suffix_format.replace("YYYY", "%Y").replace("MM", "%m").replace("DD", "%d"))
        self.log_filename = f"{log_name}-{level}{timestamp}{log_extension}"
        self.log_filepath = self.log_dir / self.log_filename
        
        # Настраиваем логирование
        self._setup_logging()
    
    def _setup_logging(self):
        """Настройка системы логирования"""
        # Определяем уровень логирования
        log_level = logging.DEBUG if self.level == "DEBUG" else logging.INFO
        
        # Настраиваем форматтер
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Настраиваем файловый хендлер
        file_handler = logging.FileHandler(self.log_filepath, encoding='utf-8')
        file_handler.setLevel(log_level)
        file_handler.setFormatter(formatter)
        
        # Настраиваем консольный хендлер
        console_handler = logging.StreamHandler()
        console_handler.setLevel(log_level)
        console_handler.setFormatter(formatter)
        
        # Настраиваем корневой логгер
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        # Очищаем хендлеры чтобы избежать дублирования
        self.logger.handlers.clear()
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
    
    def log_info(self, message):
        """Логирование информационного сообщения"""
        self.logger.info(message)
    
    def log_debug(self, message):
        """Логирование отладочного сообщения"""
        self.logger.debug(message)
    
    def log_error(self, message):
        """Логирование ошибки"""
        self.logger.error(message)
    
    def log_start(self):
        """Логирование начала работы программы"""
        self.log_info(LOG_MESSAGES["start"])
    
    def log_end(self):
        """Логирование завершения работы программы"""
        self.log_info(LOG_MESSAGES["end"])

# =============================================================================
# КЛАСС ДЛЯ СОЗДАНИЯ ТЕСТОВЫХ ДАННЫХ
# =============================================================================

class TestDataGenerator:
    """Класс для создания тестовых данных для анализа эффективности"""
    
    def __init__(self, work_dir, logger):
        """
        Инициализация генератора тестовых данных
        
        Args:
            work_dir (str): Рабочая директория
            logger (DataProcessorLogger): Объект логгера
        """
        self.work_dir = Path(work_dir)
        self.logger = logger
        self.start_time = None
        self.errors_count = 0
        self.files_created = 0
        self.employees_created = 0
        
        # Создаем необходимые директории
        self._create_directories()
        
        # Создаем распределение ГОСБ по ТБ
        self._create_tb_gosb_mapping()
    
    def _create_directories(self):
        """Создание необходимых директорий"""
        directories = [
            self.work_dir / INPUT_FOLDER,
            self.work_dir / OUTPUT_FOLDER,
            self.work_dir / LOGS_FOLDER
        ]
        
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
            self.logger.log_debug(LOG_MESSAGES["directory_ready"].format(directory))
    
    def _create_tb_gosb_mapping(self):
        """Создание распределения ГОСБ по ТБ"""
        self.tb_gosb_mapping = {}
        current_index = 0
        
        for tb in TERRITORIAL_BANKS:
            # Определяем количество ГОСБ для данного ТБ
            if tb == "Байкальский банк":
                gosb_count = 5
            elif tb == "Волго-Вятский банк":
                gosb_count = 12
            elif tb == "Дальневосточный банк":
                gosb_count = 9
            elif tb == "Московский банк":
                gosb_count = 5
            elif tb == "Поволжский банк":
                gosb_count = 8
            elif tb == "Северо-Западный банк":
                gosb_count = 10
            elif tb == "Сибирский банк":
                gosb_count = 8
            elif tb == "Среднерусский банк":
                gosb_count = 13
            elif tb == "Уральский банк":
                gosb_count = 7
            elif tb == "Центрально-Черноземный банк":
                gosb_count = 8
            elif tb == "Юго-Западный банк":
                gosb_count = 12
            else:
                gosb_count = 10
            
            # Берем ГОСБ для данного ТБ
            self.tb_gosb_mapping[tb] = HEAD_OFFICES[current_index:current_index + gosb_count]
            current_index += gosb_count
        
        self.logger.log_debug(LOG_MESSAGES["tb_mapping_created"].format(len(self.tb_gosb_mapping)))
    
    def _generate_tn(self):
        """Генерация табельного номера с настраиваемым форматом"""
        # Получаем настройки форматирования для ТН 10
        tn_format = COLUMN_SPECIAL_FORMATS.get('ТН 10', {})
        total_digits = tn_format.get('total_digits', 10)
        leading_zeros = tn_format.get('leading_zeros', 10)
        min_value = tn_format.get('min_value', 1)
        max_value = tn_format.get('max_value', 9999999999)
        
        # Генерируем случайное число в заданном диапазоне
        tn_number = np.random.randint(min_value, max_value + 1)
        
        # Форматируем с лидирующими нулями до общего количества знаков
        # Принудительно создаем строку с лидирующими нулями
        formatted_tn = str(tn_number).zfill(total_digits)
        
        # Возвращаем строку напрямую - pandas должен сохранить её как текст
        return formatted_tn
    
    def _generate_fio(self):
        """Генерация уникального ФИО"""
        # Списки имен, фамилий и отчеств
        first_names_male = [
            "Александр", "Сергей", "Владимир", "Дмитрий", "Андрей", "Алексей", "Максим", "Иван", "Михаил", "Николай",
            "Артем", "Денис", "Евгений", "Даниил", "Роман", "Тимур", "Владислав", "Павел", "Константин", "Игорь"
        ]
        
        first_names_female = [
            "Анна", "Мария", "Елена", "Ольга", "Татьяна", "Наталья", "Ирина", "Светлана", "Юлия", "Екатерина",
            "Анастасия", "Дарья", "Ксения", "Виктория", "Полина", "Алиса", "София", "Вероника", "Арина", "Диана"
        ]
        
        last_names = [
            "Иванов", "Смирнов", "Кузнецов", "Попов", "Васильев", "Петров", "Соколов", "Михайлов", "Новиков", "Федоров",
            "Морозов", "Волков", "Алексеев", "Лебедев", "Семенов", "Егоров", "Павлов", "Козлов", "Степанов", "Николаев"
        ]
        
        middle_names_male = [
            "Александрович", "Сергеевич", "Владимирович", "Дмитриевич", "Андреевич", "Алексеевич", "Максимович", "Иванович", "Михайлович", "Николаевич"
        ]
        
        middle_names_female = [
            "Александровна", "Сергеевна", "Владимировна", "Дмитриевна", "Андреевна", "Алексеевна", "Максимовна", "Ивановна", "Михайловна", "Николаевна"
        ]
        
        # Выбираем пол случайно
        is_male = np.random.choice([True, False])
        
        if is_male:
            first_name = np.random.choice(first_names_male)
            middle_name = np.random.choice(middle_names_male)
        else:
            first_name = np.random.choice(first_names_female)
            middle_name = np.random.choice(middle_names_female)
        
        last_name = np.random.choice(last_names)
        
        return f"{last_name} {first_name} {middle_name}"
    
    def _generate_effective_status(self):
        """Генерация статуса эффективности"""
        # 80% эффективных, 20% неэффективных
        is_effective = np.random.random() < DATA_PARAMS["effective_share"]
        return "👍" if is_effective else "👎"
    
    def _generate_operational_income_data(self):
        """Генерация данных об операционном доходе"""
        # Операционный доход на 31 июля 2025 (финальный)
        income_july = np.random.randint(DATA_PARAMS["operational_income_final_min"], DATA_PARAMS["operational_income_final_max"] + 1)
        
        # Операционный доход на 20 августа 2025 (текущий, обязательно >= финального)
        # Минимальный доход = доход на 31 июля, максимальный = заданный максимум
        income_august = np.random.randint(income_july, DATA_PARAMS["operational_income_current_max"] + 1)
        
        # Прирост в процентах
        growth_percent = ((income_august - income_july) / income_july * 100) if income_july > 0 else 0
        
        # Прирост в тыс. руб.
        growth_amount = income_august - income_july
        
        # ОД конец квартала = август 2025
        od_quarter = income_august
        
        return {
            'operational_income_july': income_july,
            'operational_income_august': income_august,
            'growth_percent': round(growth_percent, 2),
            'growth_amount': growth_amount,
            'od_quarter': income_august
        }
    
    def create_sample_data(self):
        """Создание тестовых данных для анализа эффективности"""
        self.start_time = time.time()
        
        try:
            self.logger.log_info(LOG_MESSAGES["data_generation_start"])
            
            # Создаем списки для данных
            data1 = []  # Данные на 31 июля 2025 года
            data2 = []  # Данные на 20 августа 2025 года
            used_fios = set()
            
            # Генерируем базовый список сотрудников для 31 июля 2025 года
            base_employees = []
            for i in range(DATA_PARAMS["total_employees"]):
                # Генерируем уникальный ФИО
                while True:
                    fio = self._generate_fio()
                    if fio not in used_fios:
                        used_fios.add(fio)
                        break
                
                # Выбираем случайный ТБ и ГОСБ
                tb = np.random.choice(TERRITORIAL_BANKS)
                gosb = np.random.choice(self.tb_gosb_mapping[tb])
                
                # Генерируем остальные данные для 31 июля 2025 года
                tn = self._generate_tn()
                effective_status = self._generate_effective_status()
                
                # Генерируем базовый доход (как бы на начало периода) и доход на 31 июля (финальный)
                # Базовый доход в диапазоне от 60% до 90% от минимального финального дохода
                base_income_min = int(DATA_PARAMS["operational_income_final_min"] * 0.6)
                base_income_max = int(DATA_PARAMS["operational_income_final_min"] * 0.9)
                base_income = np.random.randint(base_income_min, base_income_max + 1)
                income_july = np.random.randint(DATA_PARAMS["operational_income_final_min"], DATA_PARAMS["operational_income_final_max"] + 1)
                
                # Вычисляем прирост от базового дохода до 31 июля (финальный)
                growth_percent_july = ((income_july - base_income) / base_income * 100) if base_income > 0 else 0
                growth_amount_july = income_july - base_income
                
                # Создаем строку данных для 31 июля 2025 года (финальный период)
                row_july = {
                    'ТН 10': tn,
                    'ТБ': tb,
                    'ГОСБ': gosb,
                    'КМ': fio,
                    'Эффективный КМ': effective_status,
                    '2025, тыс. руб.': income_july,
                    '2024, тыс. руб. на конец месяца': base_income,
                    'Прирост, %': round(growth_percent_july, 2),
                    'Прирост, тыс. руб.': growth_amount_july,
                    'ОД конец квартала, тыс. руб.': income_july
                }
                
                data1.append(row_july)
                base_employees.append({
                    'fio': fio,
                    'tn': tn,
                    'tb': tb,
                    'gosb': gosb,
                    'effective_status': effective_status,
                    'income_july': income_july
                })
                
                # Логируем прогресс каждые 100 сотрудников
                if (i + 1) % 100 == 0:
                    self.logger.log_debug(LOG_MESSAGES["progress_employees"].format(i + 1))
            
            # Теперь создаем данные для 20 августа 2025 года
            # 90% сотрудников остаются, 5% новых, 5% убираем
            overlap_count = int(DATA_PARAMS["total_employees"] * DATA_PARAMS["employee_overlap"])
            new_count = int(DATA_PARAMS["total_employees"] * DATA_PARAMS["new_employees_share"])
            removed_count = DATA_PARAMS["total_employees"] - overlap_count - new_count
            
            # Сотрудники, которые остаются (90%)
            remaining_employees = np.random.choice(base_employees, overlap_count, replace=False)
            
            # Создаем данные для оставшихся сотрудников
            for emp in remaining_employees:
                # Генерируем доход на 20 августа (текущий, >= дохода на 31 июля)
                # Минимальный доход = доход на 31 июля, максимальный = заданный максимум
                income_august = np.random.randint(emp['income_july'], DATA_PARAMS["operational_income_current_max"] + 1)
                
                # Вычисляем прирост от финального до 31 июля до текущего на 20 августа
                growth_percent = ((income_august - emp['income_july']) / emp['income_july'] * 100) if emp['income_july'] > 0 else 0
                growth_amount = income_august - emp['income_july']
                
                row_august = {
                    'ТН 10': emp['tn'],
                    'ТБ': emp['tb'],
                    'ГОСБ': emp['gosb'],
                    'КМ': emp['fio'],
                    'Эффективный КМ': emp['effective_status'],
                    '2025, тыс. руб.': income_august,
                    '2024, тыс. руб. на конец месяца': emp['income_july'],
                    'Прирост, %': round(growth_percent, 2),
                    'Прирост, тыс. руб.': growth_amount,
                    'ОД конец квартала, тыс. руб.': income_august
                }
                
                data2.append(row_august)
            
            # Добавляем новых сотрудников (5%)
            for i in range(new_count):
                # Генерируем уникальный ФИО
                while True:
                    fio = self._generate_fio()
                    if fio not in used_fios:
                        used_fios.add(fio)
                        break
                
                # Выбираем случайный ТБ и ГОСБ
                tb = np.random.choice(TERRITORIAL_BANKS)
                gosb = np.random.choice(self.tb_gosb_mapping[tb])
                
                # Генерируем данные для нового сотрудника
                tn = self._generate_tn()
                effective_status = self._generate_effective_status()
                income_data = self._generate_operational_income_data()
                
                # Создаем строку данных для нового сотрудника
                row_august = {
                    'ТН 10': tn,
                    'ТБ': tb,
                    'ГОСБ': gosb,
                    'КМ': fio,
                    'Эффективный КМ': effective_status,
                    '2025, тыс. руб.': income_data['operational_income_august'],
                    '2024, тыс. руб. на конец месяца': income_data['operational_income_july'],
                    'Прирост, %': income_data['growth_percent'],
                    'Прирост, тыс. руб.': income_data['growth_amount'],
                    'ОД конец квартала, тыс. руб.': income_data['od_quarter']
                }
                
                data2.append(row_august)
            
            # Создаем DataFrame'ы
            df1 = pd.DataFrame(data1)  # Данные на 31 июля 2025 года
            df2 = pd.DataFrame(data2)  # Данные на 20 августа 2025 года
            self.employees_created = len(df1) + len(df2)
            
            # Анализируем распределение
            self._analyze_distribution(df1, df2)
            
            # Сохраняем файлы
            self._save_data_files(df1, df2)
            
            self.logger.log_info(LOG_MESSAGES["data_generation_end"])
            
        except Exception as e:
            error_msg = LOG_MESSAGES["generation_error"].format(str(e))
            self.logger.log_error(error_msg)
            self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
            self.errors_count += 1
        
        finally:
            # Генерируем сводку
            self._generate_summary()
    
    def _analyze_distribution(self, df1, df2):
        """Анализ распределения данных по двум файлам"""
        # Анализ файла 1 (31 июля 2025 года)
        self.logger.log_info(LOG_MESSAGES["analysis_file1"])
        tb_distribution_1 = df1['ТБ'].value_counts()
        self.logger.log_info(LOG_MESSAGES["tb_distribution"].format(dict(tb_distribution_1)))
        
        effective_distribution_1 = df1['Эффективный КМ'].value_counts()
        self.logger.log_info(LOG_MESSAGES["effective_distribution"].format(dict(effective_distribution_1)))
        
        # Анализ файла 2 (20 августа 2025 года)
        self.logger.log_info(LOG_MESSAGES["analysis_file2"])
        tb_distribution_2 = df2['ТБ'].value_counts()
        self.logger.log_info(LOG_MESSAGES["tb_distribution"].format(dict(tb_distribution_2)))
        
        effective_distribution_2 = df2['Эффективный КМ'].value_counts()
        self.logger.log_info(LOG_MESSAGES["effective_distribution"].format(dict(effective_distribution_2)))
        
        # Анализ перекрытия сотрудников
        self.logger.log_info(LOG_MESSAGES["analysis_overlap"])
        employees_july = set(df1['КМ'])
        employees_august = set(df2['КМ'])
        
        overlap_employees = employees_july.intersection(employees_august)
        new_employees = employees_august - employees_july
        removed_employees = employees_july - employees_august
        
        self.logger.log_info(LOG_MESSAGES["employees_july_count"].format(len(employees_july)))
        self.logger.log_info(LOG_MESSAGES["employees_august_count"].format(len(employees_august)))
        self.logger.log_info(LOG_MESSAGES["overlap_info_detailed"].format(len(overlap_employees), len(overlap_employees)/len(employees_july)*100))
        self.logger.log_info(LOG_MESSAGES["new_employees_info_detailed"].format(len(new_employees), len(new_employees)/len(employees_august)*100))
        self.logger.log_info(LOG_MESSAGES["removed_employees_info_detailed"].format(len(removed_employees), len(removed_employees)/len(employees_july)*100))
        
        # Проверяем уникальность ТН и ФИО в каждом файле
        unique_tn_1 = df1['ТН 10'].nunique()
        unique_fio_1 = df1['КМ'].nunique()
        unique_tn_2 = df2['ТН 10'].nunique()
        unique_fio_2 = df2['КМ'].nunique()
        
        self.logger.log_debug(LOG_MESSAGES["file1_info"].format(unique_tn_1, unique_fio_1))
        self.logger.log_debug(LOG_MESSAGES["file2_info"].format(unique_tn_2, unique_fio_2))
        
        if unique_tn_1 != len(df1):
            self.logger.log_error(LOG_MESSAGES["duplicate_tn_error"])
        if unique_fio_1 != len(df1):
            self.logger.log_error(LOG_MESSAGES["duplicate_fio_error"])
        if unique_tn_2 != len(df2):
            self.logger.log_error(LOG_MESSAGES["duplicate_tn_error"])
        if unique_fio_2 != len(df2):
            self.logger.log_error(LOG_MESSAGES["duplicate_tn_error"])
    
    def _save_data_files(self, df1, df2):
        """Сохранение данных в два файла"""
        try:
            # Формируем имена файлов с временной меткой
            timestamp = datetime.now().strftime("_%Y%m%d_%H%M%S")
            filename1 = f"data1{timestamp}.xlsx"  # Файл на 31 июля 2025 года
            filename2 = f"data2{timestamp}.xlsx"  # Файл на 20 августа 2025 года
            
            file_path1 = self.work_dir / INPUT_FOLDER / filename1
            file_path2 = self.work_dir / INPUT_FOLDER / filename2
            
            # Сохраняем файл 1 (31 июля 2025 года)
            df1.to_excel(file_path1, index=False, engine='openpyxl')
            self.logger.log_info(LOG_MESSAGES["test_file_created"].format(filename1))
            self.logger.log_debug(LOG_MESSAGES["file_saved_debug"].format(file_path1))
            self.logger.log_debug(LOG_MESSAGES["file_size_debug"].format(len(df1), len(df1.columns)))
            self.files_created += 1
            
            # Сохраняем файл 2 (20 августа 2025 года)
            df2.to_excel(file_path2, index=False, engine='openpyxl')
            self.logger.log_info(LOG_MESSAGES["test_file_created"].format(filename2))
            self.logger.log_debug(LOG_MESSAGES["file_saved_debug"].format(file_path2))
            self.logger.log_debug(LOG_MESSAGES["file_size_debug"].format(len(df2), len(df2.columns)))
            self.files_created += 1
            
        except Exception as e:
            error_msg = LOG_MESSAGES["save_error"].format(str(e))
            self.logger.log_error(error_msg)
            self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
            self.errors_count += 1
    
    def _generate_summary(self):
        """Генерация сводки создания тестовых данных"""
        end_time = time.time()
        execution_time = end_time - self.start_time
        
        summary = {
            'execution_time': format_execution_time(execution_time),
            'employees_created': self.employees_created,
            'files_created': self.files_created,
            'errors_count': self.errors_count
        }
        
        # Логируем сводку
        self.logger.log_info(LOG_MESSAGES["summary"].format(summary))
        self.logger.log_info(LOG_MESSAGES["time_elapsed"].format(format_execution_time(execution_time)))
        self.logger.log_info(LOG_MESSAGES["employees_created"].format(self.employees_created))
        self.logger.log_info(LOG_MESSAGES["outputs_created"].format(self.files_created))
        self.logger.log_info(LOG_MESSAGES["errors_count"].format(self.errors_count))
        
        return summary

# =============================================================================
# КЛАСС ДЛЯ ОБРАБОТКИ ДАННЫХ
# =============================================================================

class DataProcessor:
    """Основной класс для обработки данных"""
    
    def __init__(self, work_dir, logger):
        """
        Инициализация процессора данных
        
        Args:
            work_dir (str): Рабочая директория
            logger (DataProcessorLogger): Объект логгера
        """
        self.work_dir = Path(work_dir)
        self.logger = logger
        self.start_time = None
        self.errors_count = 0
        self.files_processed = 0
        self.outputs_created = 0
        
        # Создаем необходимые директории
        self._create_directories()
    
    def _create_directories(self):
        """Создание необходимых директорий"""
        directories = [
            self.work_dir / INPUT_FOLDER,
            self.work_dir / OUTPUT_FOLDER,
            self.work_dir / LOGS_FOLDER
        ]
        
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
            self.logger.log_debug(LOG_MESSAGES["directory_ready"].format(directory))
    
    def load_excel_files(self):
        """
        Загрузка данных из Excel файлов
        
        Returns:
            list: Список загруженных DataFrame'ов
        """
        start_time = time.time()
        dataframes = []
        
        for file_config in INPUT_FILES:
            try:
                file_path = self.work_dir / INPUT_FOLDER / f"{file_config['name']}{file_config['extension']}"
                
                if file_path.exists():
                    # Загружаем Excel файл
                    df = pd.read_excel(file_path)
                    dataframes.append({
                        'name': file_config['name'],
                        'data': df,
                        'file_path': file_path
                    })
                    
                    self.logger.log_info(LOG_MESSAGES["file_loaded"].format(file_path.name))
                    self.logger.log_debug(LOG_MESSAGES["rows_columns_loaded"].format(len(df), len(df.columns), file_path.name))
                    self.files_processed += 1
                else:
                    self.logger.log_error(LOG_MESSAGES["file_not_found"].format(file_path))
                    self.errors_count += 1
                    
            except Exception as e:
                error_msg = LOG_MESSAGES["load_file_error"].format(file_config['name'], str(e))
                self.logger.log_error(error_msg)
                self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
                self.errors_count += 1
        
        end_time = time.time()
        execution_time = end_time - start_time
        self.logger.log_debug(LOG_MESSAGES["file_loading_time"].format(format_execution_time(execution_time)))
        
        return dataframes
    
    def process_data(self, dataframes):
        """
        Обработка загруженных данных с объединением и расчетом новых колонок
        
        Args:
            dataframes (list): Список загруженных DataFrame'ов
            
        Returns:
            pd.DataFrame: Обработанные данные
        """
        start_time = time.time()
        self.logger.log_info(LOG_MESSAGES["processing_start"])
        
        if not dataframes:
            self.logger.log_error(LOG_MESSAGES["no_data_to_process"])
            return pd.DataFrame()
        
        try:
            # Находим файлы data1 и data2 по именам из INPUT_FILES
            df1 = None
            df2 = None
            
            # Получаем имена файлов из конфигурации (без расширения)
            file1_name = INPUT_FILES[0]['name']
            file2_name = INPUT_FILES[1]['name']
            
            for df_info in dataframes:
                if df_info['name'] == file1_name:
                    df1 = df_info['data']
                elif df_info['name'] == file2_name:
                    df2 = df_info['data']
            
            if df1 is None or df2 is None:
                self.logger.log_error(LOG_MESSAGES["files_not_found"].format(file1_name, file2_name))
                return pd.DataFrame()
            
            self.logger.log_debug(LOG_MESSAGES["files_loaded_info"].format(len(df1), len(df2)))
            
            # Создаем список уникальных значений ТН 10, ТБ, ГОСБ, ФИО
            # Объединяем все уникальные ТН из обоих файлов
            all_tn = pd.concat([
                df1[['ТН 10', 'ТБ', 'ГОСБ', 'КМ']].drop_duplicates(),
                df2[['ТН 10', 'ТБ', 'ГОСБ', 'КМ']].drop_duplicates()
            ]).drop_duplicates(subset=['ТН 10'], keep='last')
            
            self.logger.log_debug(LOG_MESSAGES["unique_tn_list_created"].format(len(all_tn)))
            
            # Создаем результирующий DataFrame
            result_data = []
            
            for _, row in all_tn.iterrows():
                tn = row['ТН 10']
                tb = row['ТБ']
                gosb = row['ГОСБ']
                fio = row['КМ']
                
                # Ищем данные в DataFrame
                data1_row = df1[df1['ТН 10'] == tn]
                data2_row = df2[df2['ТН 10'] == tn]
                
                # Получаем значения из файла 1
                od_current = data1_row['2025, тыс. руб.'].iloc[0] if len(data1_row) > 0 else 0
                od_previous = data1_row['2024, тыс. руб. на конец месяца'].iloc[0] if len(data1_row) > 0 else 0
                
                # Получаем эффективность из файла 2, если нет - из файла 1
                if len(data2_row) > 0:
                    effectiveness = data2_row['Эффективный КМ'].iloc[0]
                elif len(data1_row) > 0:
                    effectiveness = data1_row['Эффективный КМ'].iloc[0]
                else:
                    effectiveness = "👎"
                
                # Конвертируем эффективность в числовое значение
                effectiveness_num = 1 if effectiveness == "👍" else 0
                
                # Рассчитываем темп ОД
                if od_previous == 0:
                    if od_current > 0:
                        temp_od = 100
                    elif od_current < 0:
                        temp_od = -100
                    else:
                        temp_od = 0
                else:
                    temp_od = (od_current - od_previous) / abs(od_previous) * 100
                
                # Создаем строку результата
                result_row = {
                    'ТН 10': tn,
                    'ТБ': tb,
                    'ГОСБ': gosb,
                    'ФИО': fio,
                    'ЭФ.КМ': effectiveness_num,
                    'ОД ТЕКУЩИЙ': od_current,
                    'ранг ОД BANK': 0,  # Будет пересчитано позже
                    'ранг ОД TB': 0,    # Будет пересчитано позже
                    'ОД ПРОШЛЫЙ': od_previous,
                    'прирост': od_current - od_previous,
                    'темп': round(temp_od, 2),
                    'вып условий': 1 if (od_current - od_previous) > 0 else 0,
                    'СТРАНА 50': 0,  # Будет заполнено процентилями
                    'СТРАНА 75': 0,  # Будет заполнено процентилями
                    'СТРАНА 90': 0,  # Будет заполнено процентилями
                    'ТБ 25': 0,      # Будет заполнено процентилями
                    'ТБ 50': 0,      # Будет заполнено процентилями
                    'ТБ 75': 0,      # Будет заполнено процентилями
                    'ГОСБ 25': 0,    # Будет заполнено процентилями
                    'ГОСБ 50': 0,    # Будет заполнено процентилями
                    'ГОСБ 75': 0,    # Будет заполнено процентилями
                    'КОД вывода': 0,  # Будет рассчитано позже
                    'число страна': 0,  # Будет пересчитано позже
                    'число ТБ': 0,      # Будет пересчитано позже
                    'число подразделение': 0,  # Будет пересчитано позже
                    'вывод': '',       # Будет заполнено позже
                }
                
                result_data.append(result_row)
            
            # Создаем DataFrame
            result_df = pd.DataFrame(result_data)
            
            # Рассчитываем ранги ОД
            self.logger.log_debug("Рассчитываем ранги ОД...")
            
            # РАНГ ОД ДЛЯ УРОВНЯ BANK - точная реализация Excel формулы
            # =СЧЁТЕСЛИМН(КМР[ОД ТЕКУЩИЙ];"<"&КМР[[#Эта строка];[ОД ТЕКУЩИЙ]])/СЧЁТ(КМР[ОД ТЕКУЩИЙ])
            result_df['ранг ОД BANK'] = result_df['ОД ТЕКУЩИЙ'].apply(
                lambda x: (result_df['ОД ТЕКУЩИЙ'] < x).sum() / len(result_df) * 100
            ).round(2)
            
            # РАНГ ОД ДЛЯ УРОВНЯ TB - точная реализация Excel формулы
            # =СЧЁТЕСЛИМН(КМР[ОД ТЕКУЩИЙ];"<"&КМР[[#Эта строка];[ОД ТЕКУЩИЙ]];КМР[ТБ];КМР[[#Эта строка];[ТБ]])/СЧЁТЕСЛИМН(КМР[ТБ];КМР[[#Эта строка];[ТБ]])
            result_df['ранг ОД TB'] = result_df.apply(
                lambda row: (
                    (result_df[result_df['ТБ'] == row['ТБ']]['ОД ТЕКУЩИЙ'] < row['ОД ТЕКУЩИЙ']).sum() /
                    len(result_df[result_df['ТБ'] == row['ТБ']]) * 100
                ) if len(result_df[result_df['ТБ'] == row['ТБ']]) > 0 else 0, axis=1
            ).round(2)
            
            # Рассчитываем процентили для трех уровней
            self.logger.log_debug("Рассчитываем процентили...")
            
            # Процентили для уровня БАНК (по всем данным)
            bank_percentiles = result_df['ОД ТЕКУЩИЙ'].quantile([p/100 for p in PERCENTILES])
            
            # Процентили для уровня ТБ (по каждому ТБ отдельно)
            tb_percentiles = {}
            for tb in result_df['ТБ'].unique():
                tb_mask = result_df['ТБ'] == tb
                tb_percentiles[tb] = result_df.loc[tb_mask, 'ОД ТЕКУЩИЙ'].quantile([p/100 for p in PERCENTILES])
            
            # Процентили для уровня ГОСБ (по каждому ГОСБ отдельно)
            gosb_percentiles = {}
            for gosb in result_df['ГОСБ'].unique():
                gosb_mask = result_df['ГОСБ'] == gosb
                gosb_percentiles[gosb] = result_df.loc[gosb_mask, 'ОД ТЕКУЩИЙ'].quantile([p/100 for p in PERCENTILES])
            
            # Заполняем колонки с процентилями
            result_df['СТРАНА 50'] = bank_percentiles[0.50]
            result_df['СТРАНА 75'] = bank_percentiles[0.75]
            result_df['СТРАНА 90'] = bank_percentiles[0.90] if 0.90 in bank_percentiles.index else bank_percentiles[0.75]
            
            # Процентили ТБ для каждой строки
            result_df['ТБ 25'] = result_df['ТБ'].map(lambda x: tb_percentiles[x][0.25])
            result_df['ТБ 50'] = result_df['ТБ'].map(lambda x: tb_percentiles[x][0.50])
            result_df['ТБ 75'] = result_df['ТБ'].map(lambda x: tb_percentiles[x][0.75])
            
            # Процентили ГОСБ для каждой строки
            result_df['ГОСБ 25'] = result_df['ГОСБ'].map(lambda x: gosb_percentiles[x][0.25])
            result_df['ГОСБ 50'] = result_df['ГОСБ'].map(lambda x: gosb_percentiles[x][0.50])
            result_df['ГОСБ 75'] = result_df['ГОСБ'].map(lambda x: gosb_percentiles[x][0.75])
            
            # Рассчитываем колонки "число страна", "число ТБ", "число подразделение"
            self.logger.log_debug("Рассчитываем ранжирование по процентилям...")
            
            # число страна - ранжирование по темпу среди всех
            result_df['число страна'] = result_df['темп'].rank(method='min', ascending=False)
            
            # число ТБ - ранжирование по темпу в рамках ТБ
            result_df['число ТБ'] = result_df.groupby('ТБ')['темп'].rank(method='min', ascending=False)
            
            # число подразделение - ранжирование по темпу в рамках ГОСБ
            result_df['число подразделение'] = result_df.groupby('ГОСБ')['темп'].rank(method='min', ascending=False)
            
            # Рассчитываем колонку "КОД вывода" согласно логике из Excel файла
            def calculate_kod_vyvoda(row):
                """
                Расчет кода вывода согласно логике из листа 't' Excel файла:
                - Код 6: выше, чем у 90% КМ в стране
                - Код 5: выше, чем у 75% КМ в стране  
                - Код 4: выше, чем у 75% КМ в тербанке (среди эффективных)
                - Код 3: выше, чем у 75% КМ в тербанке (среди всех)
                - Код 2: выше, чем у 75% КМ в ГОСБ/аппарате (среди эффективных, положительный прирост)
                - Код 1: ниже, чем у 75% КМ в ГОСБ/аппарате (среди всех, положительный прирост)
                """
                # Получаем значения для расчета
                number_strana = row['число страна']
                number_tb = row['число ТБ']
                number_gosb = row['число подразделение']
                effectiveness = row['ЭФ.КМ']  # 1 для эффективных, 0 для неэффективных
                prir = row['прирост']  # прирост ОД
                
                # Код 6: выше, чем у 90% КМ в стране
                if number_strana <= 0.1 * len(result_df):  # Топ-10% по стране
                    return 6
                
                # Код 5: выше, чем у 75% КМ в стране
                elif number_strana <= 0.25 * len(result_df):  # Топ-25% по стране
                    return 5
                
                # Код 4: выше, чем у 75% КМ в тербанке (среди эффективных)
                elif effectiveness == 1 and number_tb <= 0.25 * len(result_df[result_df['ТБ'] == row['ТБ']]):
                    return 4
                
                # Код 3: выше, чем у 75% КМ в тербанке (среди всех)
                elif number_tb <= 0.25 * len(result_df[result_df['ТБ'] == row['ТБ']]):
                    return 3
                
                # Код 2: выше, чем у 75% КМ в ГОСБ/аппарате (среди эффективных, положительный прирост)
                elif effectiveness == 1 and prir > 0 and number_gosb <= 0.25 * len(result_df[result_df['ГОСБ'] == row['ГОСБ']]):
                    return 2
                
                # Код 1: ниже, чем у 75% КМ в ГОСБ/аппарате (среди всех, положительный прирост)
                elif prir > 0 and number_gosb <= 0.25 * len(result_df[result_df['ГОСБ'] == row['ГОСБ']]):
                    return 1
                
                # Если не подходит ни под один код
                else:
                    return 0
            
            result_df['КОД вывода'] = result_df.apply(calculate_kod_vyvoda, axis=1)
            
            # Рассчитываем колонку "вывод" на основе кода вывода
            def calculate_output_by_code(row):
                """
                Возвращает текстовое описание результата согласно коду вывода
                """
                kod = row['КОД вывода']
                
                if kod == 6:
                    return "выше, чем у 90% КМ в стране"
                elif kod == 5:
                    return "выше, чем у 75% КМ в стране"
                elif kod == 4:
                    return "выше, чем у 75% КМ в тербанке (среди эффективных)"
                elif kod == 3:
                    return "выше, чем у 75% КМ в тербанке"
                elif kod == 2:
                    return "выше, чем у 75% КМ в ГОСБ/аппарате (среди эффективных)"
                elif kod == 1:
                    return "ниже, чем у 75% КМ в ГОСБ/аппарате"
                else:
                    return "обычный результат"
            
            result_df['вывод'] = result_df.apply(calculate_output_by_code, axis=1)
            
            end_time = time.time()
            execution_time = end_time - start_time
            self.logger.log_debug(LOG_MESSAGES["data_processing_time"].format(format_execution_time(execution_time)))
            self.logger.log_debug(LOG_MESSAGES["data_processed_info"].format(len(result_df), len(result_df.columns)))
            self.logger.log_info(LOG_MESSAGES["processing_end"])
            
            return result_df
            
        except Exception as e:
            error_msg = LOG_MESSAGES["processing_error"].format(str(e))
            self.logger.log_error(error_msg)
            self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
            self.errors_count += 1
            return pd.DataFrame()
    
    def save_outputs(self, processed_data):
        """
        Сохранение обработанных данных в выходные файлы
        
        Args:
            processed_data (pd.DataFrame): Обработанные данные
        """
        start_time = time.time()
        if processed_data.empty:
            self.logger.log_error(LOG_MESSAGES["no_data_to_save"])
            return
        
        for output_config in OUTPUT_FILES:
            try:
                # Формируем имя файла с суффиксом
                timestamp = datetime.now().strftime(
                    output_config["suffix_format"]
                    .replace("YYYY", "%Y")
                    .replace("MM", "%m")
                    .replace("DD", "%d")
                    .replace("HH", "%H")
                    .replace("MM", "%M")
                    .replace("SS", "%S")
                )
                
                filename = f"{output_config['name']}{timestamp}{output_config['extension']}"
                file_path = self.work_dir / OUTPUT_FOLDER / filename
                
                # Сохраняем файл в зависимости от формата
                # ЗАКЛАДКА ДЛЯ БУДУЩЕГО CSV:
                # if output_config['extension'].lower() == '.csv':
                #     # Сохраняем CSV с разделителем ";"
                #     processed_data.to_csv(file_path, sep=';', index=False, encoding='utf-8')
                # if output_config['extension'].lower() == '.xlsx':
                # 
                # Текущая реализация - только Excel файл
                if output_config['extension'].lower() == '.xlsx':
                    # Сохраняем Excel с автофильтром и форматированием
                    processed_data.to_excel(file_path, index=False, engine='openpyxl')
                    
                    # Открываем файл для добавления автофильтра и форматирования
                    
                    # Загружаем сохраненный файл
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # Добавляем автофильтр начиная с колонки E (5-я колонка) и строки 2
                    # Определяем диапазон для автофильтра (все данные)
                    max_row = len(processed_data) + 1  # +1 потому что pandas.to_excel добавляет заголовки
                    max_col = len(processed_data.columns)
                    
                    # Функция для получения буквы колонки по номеру
                    def get_column_letter(col_num):
                        """Преобразует номер колонки в букву Excel (A, B, C, ..., Z, AA, AB, ...)"""
                        result = ""
                        while col_num > 0:
                            col_num, remainder = divmod(col_num - 1, 26)
                            result = chr(65 + remainder) + result
                        return result
                    
                    # Устанавливаем автофильтр на диапазон A1:последняя_колонка_последняя_строка (заголовки + данные)
                    # A1 = заголовки, A2:последняя_колонка_последняя_строка = данные
                    last_col_letter = get_column_letter(max_col)
                    ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"
                    
                    # Фиксируем панели на уровне A2 (колонка A, строка 2) - заголовки остаются видимыми
                    ws.freeze_panes = "A2"
                    
                    # Функция для получения настроек форматирования колонки
                    def get_column_format_config(column_name):
                        """Получает настройки форматирования для колонки из групп или специальных настроек"""
                        # Сначала проверяем специальные настройки (переопределяют групповые)
                        if column_name in COLUMN_SPECIAL_FORMATS:
                            special_config = COLUMN_SPECIAL_FORMATS[column_name].copy()
                            
                            # Ищем группу для этой колонки
                            for group_name, group_config in COLUMN_FORMAT_GROUPS.items():
                                if column_name in group_config['columns']:
                                    # Объединяем групповые настройки со специальными
                                    format_config = group_config.copy()
                                    format_config.update(special_config)  # Специальные переопределяют групповые
                                    return format_config
                            
                            # Если колонка не найдена в группах, возвращаем только специальные
                            return special_config
                        
                        # Если нет специальных настроек, ищем в группах
                        for group_name, group_config in COLUMN_FORMAT_GROUPS.items():
                            if column_name in group_config['columns']:
                                return group_config.copy()
                        
                        # Если колонка не найдена нигде, возвращаем None
                        return None
                    
                    # Применяем форматирование колонок согласно групповой системе
                    for col in range(1, max_col + 1):
                        column_letter = get_column_letter(col)
                        column_name = processed_data.columns[col - 1]  # Получаем название колонки
                        
                        # Получаем настройки форматирования для колонки
                        format_config = get_column_format_config(column_name)
                        
                        if format_config:
                            # Устанавливаем ширину колонки
                            ws.column_dimensions[column_letter].width = format_config.get('width', 15)
                            
                            # Применяем форматирование ко всем ячейкам в колонке (кроме заголовка)
                            for row in range(2, max_row + 1):  # Начинаем со 2-й строки (после заголовка)
                                cell = ws[f"{column_letter}{row}"]
                                
                                # Применяем специальное форматирование для padded_number
                                if format_config.get('format_type') == 'padded_number':
                                    # Для чисел с лидирующими нулями устанавливаем текстовый формат
                                    cell.number_format = '@'  # Текстовый формат Excel
                                    # Применяем выравнивание по левому краю для лучшей читаемости
                                    cell.alignment = Alignment(horizontal='left', vertical='center')
                                else:
                                    # Применяем числовой формат
                                    if format_config.get('format') == 'number' and 'number_format' in format_config:
                                        cell.number_format = format_config['number_format']
                                    
                                    # Применяем выравнивание
                                    alignment = format_config.get('alignment', 'left')
                                    if alignment == 'center':
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                    elif alignment == 'right':
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                    elif alignment == 'left':
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            # Для колонок без настроек - форматируем по содержимому
                            max_width = 0
                            for row in range(1, max_row + 1):
                                cell_value = ws[f"{column_letter}{row}"].value
                                if cell_value is not None:
                                    if isinstance(cell_value, (int, float)):
                                        width = len(str(cell_value)) + 2
                                    else:
                                        width = len(str(cell_value)) + 1
                                    max_width = max(max_width, width)
                            
                            ws.column_dimensions[column_letter].width = min(max_width + 2, 50)
                    
                    # Сохраняем изменения
                    wb.save(file_path)
                    wb.close()
                    
                    self.logger.log_debug(LOG_MESSAGES["autofilter_added"].format(f"A1:{last_col_letter}{max_row}"))
                    self.logger.log_debug(LOG_MESSAGES["panes_frozen"])
                    # Логируем информацию о примененном форматировании
                    formatted_columns = 0
                    for group_name, group_config in COLUMN_FORMAT_GROUPS.items():
                        formatted_columns += len(group_config['columns'])
                    
                    # Логируем информацию о примененном форматировании
                    formatted_columns = 0
                    for group_name, group_config in COLUMN_FORMAT_GROUPS.items():
                        formatted_columns += len(group_config['columns'])
                    
                    # Подсчитываем колонки со специальным форматированием
                    special_formatted = 0
                    for col_name, col_config in COLUMN_SPECIAL_FORMATS.items():
                        if col_config.get('format_type') == 'padded_number':
                            special_formatted += 1
                    
                    self.logger.log_debug(LOG_MESSAGES["group_formatting_applied"].format(len(COLUMN_FORMAT_GROUPS), formatted_columns))
                    self.logger.log_debug(LOG_MESSAGES["special_formats_applied"].format(len(COLUMN_SPECIAL_FORMATS)))
                    if special_formatted > 0:
                        self.logger.log_debug(LOG_MESSAGES["padded_number_formatted"].format(special_formatted))
                
                self.logger.log_info(LOG_MESSAGES["file_saved"].format(filename))
                self.logger.log_debug(LOG_MESSAGES["file_saved_debug_old"].format(file_path))
                self.outputs_created += 1
                
            except Exception as e:
                error_msg = LOG_MESSAGES["save_error"].format(f"{output_config['name']}: {str(e)}")
                self.logger.log_error(error_msg)
                self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
                self.errors_count += 1
        
        end_time = time.time()
        execution_time = end_time - start_time
        self.logger.log_debug(LOG_MESSAGES["file_saving_time"].format(format_execution_time(execution_time)))
    
    def generate_summary(self):
        """Генерация сводки выполнения программы"""
        end_time = time.time()
        execution_time = end_time - self.start_time
        
        summary = {
            'execution_time': format_execution_time(execution_time),
            'files_processed': self.files_processed,
            'outputs_created': self.outputs_created,
            'errors_count': self.errors_count
        }
        
        # Логируем сводку
        self.logger.log_info(LOG_MESSAGES["summary"].format(summary))
        self.logger.log_info(LOG_MESSAGES["time_elapsed"].format(format_execution_time(execution_time)))
        self.logger.log_info(LOG_MESSAGES["files_processed"].format(self.files_processed))
        self.logger.log_info(LOG_MESSAGES["outputs_created"].format(self.outputs_created))
        self.logger.log_info(LOG_MESSAGES["errors_count"].format(self.errors_count))
        
        return summary
    
    def run(self):
        """Основной метод запуска обработки данных"""
        self.start_time = time.time()
        
        try:
            # Загружаем данные
            dataframes = self.load_excel_files()
            
            # Обрабатываем данные
            processed_data = self.process_data(dataframes)
            
            # Сохраняем результаты
            self.save_outputs(processed_data)
            
        except Exception as e:
            error_msg = LOG_MESSAGES["processing_error"].format(str(e))
            self.logger.log_error(error_msg)
            self.logger.log_debug(LOG_MESSAGES["details_error"].format(traceback.format_exc()))
            self.errors_count += 1
        
        finally:
            # Генерируем сводку
            self.generate_summary()

# =============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# =============================================================================

def main():
    """Главная функция программы"""
    
    try:
        # Создаем логгер
        logger = DataProcessorLogger(
            log_dir=Path(WORK_DIR) / LOGS_FOLDER,
            log_name=LOG_FILE["name"],
            log_extension=LOG_FILE["extension"],
            suffix_format=LOG_FILE["suffix_format"],
            level=LOG_LEVEL
        )
        
        # Логируем начало работы
        logger.log_start()
        
        if PROGRAM_MODE == 'create-test':
            # Режим создания тестовых данных
            logger.log_info(LOG_MESSAGES["mode_create_test"])
            generator = TestDataGenerator(WORK_DIR, logger)
            generator.create_sample_data()
            print(LOG_MESSAGES["test_data_success"])
            
        else:
            # Режим обработки данных (по умолчанию)
            logger.log_info(LOG_MESSAGES["mode_process"])
            processor = DataProcessor(WORK_DIR, logger)
            processor.run()
            print(LOG_MESSAGES["process_success"])
        
        # Логируем завершение работы
        logger.log_end()
        
    except Exception as e:
        print(LOG_MESSAGES["main_critical_error"].format(str(e)))
        sys.exit(1)

# =============================================================================
# ТОЧКА ВХОДА
# =============================================================================

if __name__ == "__main__":
    main()
